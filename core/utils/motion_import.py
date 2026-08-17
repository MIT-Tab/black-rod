import csv
from datetime import date, datetime
from io import BytesIO, StringIO

from django.db import transaction
from openpyxl import load_workbook

from core.models import Motion, Tournament


HEADER_ALIASES = {
    "motion": "motion_text",
    "text": "motion_text",
    "motion text": "motion_text",
    "background": "background_slide",
    "background slide": "background_slide",
    "date": "date_set",
    "date set": "date_set",
    "tournament set": "tournament",
    "topics": "tags",
    "topic tags": "tags",
}


def _normalized_header(value):
    header = str(value or "").strip().lower().replace("_", " ")
    return HEADER_ALIASES.get(header, header.replace(" ", "_"))


def read_motion_spreadsheet(upload):
    extension = upload.name.rsplit(".", 1)[-1].lower()
    if extension == "csv":
        text = upload.read().decode("utf-8-sig")
        raw_rows = list(csv.reader(StringIO(text)))
    else:
        workbook = load_workbook(BytesIO(upload.read()), read_only=True, data_only=True)
        raw_rows = list(workbook.active.iter_rows(values_only=True))

    if not raw_rows:
        raise ValueError("The spreadsheet is empty.")
    headers = [_normalized_header(value) for value in raw_rows[0]]
    if "motion_text" not in headers:
        raise ValueError("The spreadsheet needs a 'motion_text' column.")

    rows = []
    for row_number, values in enumerate(raw_rows[1:], start=2):
        row = {header: values[index] if index < len(values) else "" for index, header in enumerate(headers)}
        if not any(value not in (None, "") for value in row.values()):
            continue
        rows.append(validate_motion_row(row_number, row))
    return rows


def _parse_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for date_format in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(str(value).strip(), date_format).date()
        except ValueError:
            pass
    raise ValueError("use YYYY-MM-DD or MM/DD/YYYY")


def _resolve_tournament(value):
    if not value:
        return None
    raw = str(value).strip()
    if raw.isdigit():
        tournament = Tournament.objects.filter(pk=int(raw)).first()
        if tournament:
            return tournament
    matches = Tournament.objects.filter(name__iexact=raw)
    if matches.count() == 1:
        return matches.first()
    if not matches.exists():
        raise ValueError(f"tournament '{raw}' was not found")
    raise ValueError(f"tournament '{raw}' is ambiguous; use its numeric ID")


def validate_motion_row(row_number, row):
    errors = []
    text = str(row.get("motion_text") or "").strip()
    if not text:
        errors.append("motion text is required")
    try:
        date_set = _parse_date(row.get("date_set"))
    except ValueError as exc:
        date_set = None
        errors.append(f"invalid date ({exc})")
    try:
        tournament = _resolve_tournament(row.get("tournament"))
    except ValueError as exc:
        tournament = None
        errors.append(str(exc))
    raw_tags = str(row.get("tags") or "")
    tags = [tag.strip() for tag in raw_tags.replace(";", ",").split(",") if tag.strip()]
    duplicate = bool(text and Motion.objects.filter(text__iexact=text).exists())
    return {
        "row_number": row_number,
        "text": text,
        "background_slide": str(row.get("background_slide") or "").strip(),
        "date_set": date_set.isoformat() if date_set else "",
        "tournament_id": tournament.pk if tournament else None,
        "tournament": str(tournament) if tournament else "",
        "tags": tags,
        "duplicate": duplicate,
        "errors": errors,
    }


@transaction.atomic
def import_motion_rows(rows):
    created = skipped = 0
    for row in rows:
        if row["errors"] or row["duplicate"]:
            skipped += 1
            continue
        motion, was_created = Motion.objects.get_or_create(
            text=row["text"],
            defaults={
                "background_slide": row["background_slide"],
                "date_set": row["date_set"] or None,
                "tournament_id": row["tournament_id"],
            },
        )
        if not was_created:
            skipped += 1
            continue
        motion.tags.add(*row["tags"])
        created += 1
    return created, skipped
